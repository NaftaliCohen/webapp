import pandas as pd
import os
import sys
import tkinter as tk
from tkinter import messagebox
import win32com.client
import re
import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from PIL import Image, ImageTk
import webbrowser
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt


# ========== הגדרת משתנים ==========
EXCEL_DIR = r"\\filesrv\Civan Folder\Supply Chain\Procurement\יבוא ויצוא\מעקב אספקות\python_app"
os.makedirs(EXCEL_DIR, exist_ok=True)
FRAME_WIDTH, FRAME_HEIGHT = 500, 500
BG_WIDTH, BG_HEIGHT = 150 , 150


# ========== חיבור לבסיס נתונים ==========
def fetch_data():
    conn_str = (
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=xxxx;DATABASE=Civan;UID=xxxx;PWD=xxxxxx;'
    )
    conn = pyodbc.connect(conn_str)
    query = """
        SELECT  
            O1.DocNum AS 'Order Number',
            O1.CardName AS 'Supplier name',
            O2.lineNum + 1 AS 'Line Number',
            O2.ItemCode AS 'PN',
            O2.VendorNum AS 'Manufacturer PN',
            O2.Dscription AS 'Description',
            cast(O2.Quantity as int) AS 'Order quantity',
            cast(O2.OpenCreQty as int) AS 'Open quantity',
            convert(date ,O2.U_AppActDelDate) AS 'Confirmed delivery date',
            NULL AS 'Update delivery date',
            O2.U_INS_QUT_LineMemo AS 'Notes',
            O1.DocDate AS 'Date of sending Email',
            O2.U_ship_date_civan AS 'Delivery date in Civan',
            CASE
                WHEN O2.U_OrderStatus = 2 THEN 'Delivery date confirmed'
                WHEN O2.U_OrderStatus = 5 THEN 'Shipped'
                ELSE 'Promote'
            END AS Order_Status,
            O2.Project AS ' Project ',
            P.PrjName AS 'Project_Name',
            O2.U_User_category AS 'Category',
            B1.SlpName AS 'Buyer',
            CASE
                WHEN CAST(GETDATE() AS DATE) < CAST(O2.U_ship_date_civan AS DATE) THEN 'Open'
                WHEN CAST(GETDATE() AS DATE) = CAST(O2.U_ship_date_civan AS DATE) THEN 'Open'
                WHEN CAST(GETDATE() AS DATE) > CAST(O2.U_ship_date_civan AS DATE) THEN 'Late'
                ELSE 'No Date'
            END AS Orderstatus,
            C.E_MailL AS 'Email',
            O2.U_owner AS 'Demander',
            O2.U_INS_QUT_LineMemo AS 'Note'
FROM OPOR O1
        INNER JOIN POR1 O2 ON O1.DocEntry = O2.DocEntry
        LEFT JOIN OCPR C ON O1.CardCode = C.CardCode AND C.Name = 'promote orders'
        INNER JOIN OSLP B1 ON B1.SlpCode = O1.SlpCode
        INNER JOIN OPRJ P ON P.PrjCode = O2.Project
WHERE
            O1.DocDate BETWEEN '2024-01-01' AND GETDATE()
            AND O1.Docstatus = 'O'
            AND O2.Linestatus = 'O'
            AND O1.U_CoverOrder IS NULL
            AND O1.DocNum NOT LIKE '%700000%'
            AND O1.DocNum NOT LIKE '%5007215%'
ORDER BY O1.DocNum ASC, O2.lineNum +1 ASC
 """
    df = pd.read_sql(query, conn)
    conn.close()
    return df

# ========== שמירת הקובץ ==========
def save_main_excel(df):
    output_path = os.path.join(EXCEL_DIR, 'order_orders.xlsx')
    df.to_excel(output_path, index=False, engine='openpyxl')
    return output_path

# ========== שליחת מיילים ==========
def send_emails(filtered_df):
    for email in filtered_df['Email'].unique():
        supplier_df = filtered_df[filtered_df['Email'] == email]
        supplier_name = supplier_df['Supplier name'].iloc[0]
        safe_name = re.sub(r'[\\/*?:"<>|]', '_', supplier_name)
        supplier_file = os.path.join(EXCEL_DIR, f"{safe_name}.xlsx")

        supplier_df_filtered = supplier_df[[
            'Order Number', 'Supplier name', 'Line Number', 'PN', 'Manufacturer PN',
            'Description', 'Order quantity', 'Open quantity', 'Confirmed delivery date',
            'Update delivery date'
        ]]
        supplier_df_filtered.to_excel(supplier_file, index=False, engine='openpyxl')
        format_excel(supplier_file)

        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = email
        mail.Subject = f"Delivery Dates"
        mail.SentOnBehalfOfName = "Export-Import@civanlasers.com"
        mail.Body = (
            f"Hi  {supplier_name},\n\n"
            "Attached is a file with details of open purchase orders by Civan system.\n"
            "Please update dates for all orders in the file,and send the file back with dates.\n\n"
            "Kind regards,\n Procurement Team Civan"
        )
        mail.Attachments.Add(supplier_file)
        mail.Display() #send()

    messagebox.showinfo("הצלחה", "המיילים הוכנו (הוצגו ב-Outlook)")

# ========== עיצוב קובץ אקסל ==========
def format_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    row_fill_alt = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = row_fill_alt

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    wb.save(filepath)

# ========== הצגת בחירת ספקים ==========
def show_supplier_selector(data, title):
    bg_image = Image.open("background.png")
    bg_width, bg_height = bg_image.size

    window = tk.Toplevel(root)
    window.title(title)
    window.geometry(f"{bg_width}x{bg_height}")
    window.resizable(False, False)

    bg_photo = ImageTk.PhotoImage(bg_image)
    background_label = tk.Label(window, image=bg_photo)
    background_label.image = bg_photo
    background_label.place(relwidth=1, relheight=1)

    content_frame = tk.Frame(window, bg='white', width=1000, height=600)
    content_frame.place(relx=0.5, rely=0.5, anchor='center')
    content_frame.pack_propagate(False)

    canvas = tk.Canvas(content_frame, bg='white', highlightthickness=0)
    scrollbar = tk.Scrollbar(content_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas, bg='white')

    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    check_vars = []
    # חישוב כמות השורות הפתוחות לכל ספק
    supplier_counts = data.groupby('Supplier name').size().reset_index(name='Open Rows')

    # שילוב הכמות ב־unique_suppliers
    unique_suppliers = data[['Supplier name', 'Email', 'Note', ' Project ']].drop_duplicates()
    unique_suppliers = unique_suppliers.merge(supplier_counts, on='Supplier name', how='left')

    # מיון לפי שם הספק
    unique_suppliers = unique_suppliers.sort_values(by='Supplier name')

    select_all_var = tk.BooleanVar()

    def toggle_all():
        for var, _ in check_vars:
            var.set(select_all_var.get())

    # תיבת סימון Select All
    select_all_chk = tk.Checkbutton(scrollable_frame, text="Select All", variable=select_all_var,
                                    command=toggle_all, bg='white', font=('Calibri', 16, 'bold'),
                                    anchor='w', justify='left')
    select_all_chk.pack(fill="x", padx=5, pady=5, anchor="w")

    for _, row in unique_suppliers.iterrows():
        var = tk.BooleanVar()

        note = f"Note: {row['Note']}" if pd.notna(row['Note']) else ""
        project_text = f"Project: {row[' Project ']}"
        open_rows_text = f"Rows: {row['Open Rows']}"

        chk_text = f"{row['Supplier name']}  ({row['Email']}) {project_text} {note} {open_rows_text}"

        chk = tk.Checkbutton(
            scrollable_frame,
            text=chk_text,
            variable=var,
            bg='white',
            anchor='w',
            justify='left',
            wraplength=600,
            font=('Calibri', 16)
        )
        chk.pack(fill="x", padx=5, pady=2, anchor="w")
        check_vars.append((var, row['Email']))


    def send_selected():
        selected_emails = [email for var, email in check_vars if var.get()]
        if not selected_emails:
            messagebox.showwarning("שגיאה", "בחר לפחות ספק אחד")
            return
        send_emails(data[data['Email'].isin(selected_emails)])
        window.destroy()

    send_img = ImageTk.PhotoImage(Image.open("send.png").resize((150, 150)))
    send_btn = tk.Button(content_frame, image=send_img, command=send_selected, borderwidth=0, bg='white')
    send_btn.image = send_img
    send_btn.pack(pady=10)



# ========== פקודות סינון ==========
def send_open_orders():
    filtered = df[df['Orderstatus'] == 'Open']
    show_supplier_selector(filtered, "Suppliers with an open order date")

def send_late_orders():
    filtered = df[df['Orderstatus'] == 'Late']
    show_supplier_selector(filtered, "Suppliers with a late order date")

def send_no_date_orders():
    filtered = df[df['Orderstatus'] == 'No Date']
    show_supplier_selector(filtered, "Suppliers without date")

def send_all_orders():
    show_supplier_selector(df, "All Suppliers ")

# ========== רענון ==========
def refresh_app():
    root.destroy()       # סוגר את הממשק הקיים
    main()               # מפעיל אותו מחדש





# ========== פתיחת דשבורד ==========
def open_powerbi_dashboard():
    webbrowser.open_new_tab("https://app.powerbi.com/links/WMz5AUp38t?...")

def open_powerbi_dashboard_1():
    webbrowser.open_new_tab("https://app.powerbi.com/links/SpdUb05X27?...")

def create_pie_chart(parent_frame, df):
    # הכנה של הגרף
    status_counts = df['Orderstatus'].value_counts()
    labels = status_counts.index.tolist()
    sizes = status_counts.values.tolist()
    colors = ['#4CD964','#FF5252', '#FFEB3B'] # Open, No Date, Late

    fig, ax = plt.subplots(figsize=(4, 4), dpi=100)
    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=[f"{label} ({count})" for label, count in zip(labels, sizes)],
        autopct='%1.1f%%',
        colors=colors,
        textprops={'color': 'black', 'fontsize': 9, 'fontname': 'Calibri'}

    )
    ax.set_title("Order Status", fontsize=14, color='indigo',fontname = 'Calibri')
    #ax.legend(wedges, labels, loc="center left", bbox_to_anchor=(1, 0, 0.5, 1), fontsize=10)

    # הוספת הגרף לתוך המסך הראשי
    canvas = FigureCanvasTkAgg(fig, master=parent_frame)
    canvas.draw()
    widget = canvas.get_tk_widget()
    widget.grid(row=0, column=0, padx=10, pady=10, sticky="nw")


def create_project_bar_chart(parent_frame, df):
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import mplcursors

    # סופרים את מספר השורות לכל פרויקט ומחזירים את 10 הפרויקטים עם הכי הרבה מופעים
    top_projects = df[' Project '].value_counts().nlargest(10)

    # יוצרים מיפוי בין שם הפרויקט המקורי לבין Project_Name
    project_name_map = df.drop_duplicates(subset=[' Project ']).set_index(' Project ')['Project_Name'].to_dict()

    # בניית הגרף
    fig, ax = plt.subplots(figsize=(4, 4), dpi=100)
    bars = ax.bar(top_projects.index, top_projects.values, color='skyblue')

    ax.set_title("Top 10 Projects by Order Count", fontsize=14, color='indigo', fontname='Calibri')
    ax.tick_params(axis='x', rotation=45)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['bottom'].set_linewidth(1.5)

    # הצגת המספרים מעל כל עמודה
    for bar in bars:
        height = bar.get_height()
        ax.annotate(f'{height}',
                    xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3),
                    textcoords="offset points",
                    ha='center', va='bottom', fontsize=9)

    # הוספת tooltip עם mplcursors
    from bidi.algorithm import get_display
    cursor = mplcursors.cursor(bars, hover=True)
    @cursor.connect("add")
    def on_hover(sel):
        index = sel.index
        project_id = top_projects.index[index]

        # מביא את ה-Project_Name המקורי מה-DataFrame
        project_name = df.loc[df[' Project '] == project_id, 'Project_Name'].iloc[0]

        # הופך טקסט עברי לתצוגה נכונה
        project_name_display = get_display(project_name)

        sel.annotation.set_text(project_name_display)
        sel.annotation.get_bbox_patch().set(fc="white", alpha=0.9)

    # הוספת הגרף למסך
    canvas = FigureCanvasTkAgg(fig, master=parent_frame)
    canvas.draw()
    widget = canvas.get_tk_widget()
    widget.grid(row=1, column=0, padx=10, pady=10, sticky="nw")



# ========== ממשק ראשי ==========
def main():
    global root, df
    df = fetch_data()
    save_main_excel(df)

    root = tk.Tk()
    root.title("Order Management App")

    # רקע
    bg_image = Image.open("background.png")
    root.geometry(f"{bg_image.size[0]}x{bg_image.size[1]}")
    bg_photo = ImageTk.PhotoImage(bg_image)
    tk.Label(root, image=bg_photo).place(relwidth=1, relheight=1)
    root.bg_photo = bg_photo

    # מסגרת מרכזית
    frame = tk.Frame(root, width=FRAME_WIDTH, height=FRAME_HEIGHT, bg='white')
    frame.place(relx=0.5, rely=0.5, anchor='center')
    frame.pack_propagate(False)

    # --- NEW: מסגרת צדדית שמכילה גם את הגרף וגם את הכפתורים זה לצד זה ---
    content_frame = tk.Frame(frame, bg='white')
    content_frame.pack(expand=True, fill="both")


    # גרף פאי - מוצב בפינה השמאלית העליונה של root
    pie_frame = tk.Frame(root, bg='white')
    pie_frame.place(relx=0.0, rely=0.0, anchor="nw", x=10, y=10)
    create_pie_chart(pie_frame, df)

    # פריים עבור גרף העמודות בצד ימין למטה
    bar_frame = tk.Frame(root, bg='white')
    bar_frame.place(relx=1.0, rely=1.0, anchor="se", x=-10, y=-10)  # ימין למטה
    create_project_bar_chart(bar_frame, df)

    # כפתורים - עמודה ימנית
    buttons_frame = tk.Frame(content_frame, bg='white')
    buttons_frame.grid(row=0, column=1, padx=100, pady=10 ,sticky="ne")

    def load_btn_image(file):
        return ImageTk.PhotoImage(Image.open(file).resize((BG_WIDTH, BG_HEIGHT)))

    btns = [
        (load_btn_image("openorders.png"), send_open_orders),
        (load_btn_image("lateorders.png"), send_late_orders),
        (load_btn_image("nodateorders.png"), send_no_date_orders),
        (load_btn_image("allorders.png"), send_all_orders),
        (load_btn_image("powerbi_logo.png"), open_powerbi_dashboard),
        (load_btn_image("download_powerbi.jpeg"), open_powerbi_dashboard_1)
    ]

    for i, (img, cmd) in enumerate(btns):
        btn = tk.Button(buttons_frame, image=img, command=cmd, bg='white', borderwidth=0)
        btn.image = img
        btn.grid(row=i//2, column=i%2, padx=5, pady=5)

    # כפתור רענון בפינה ימנית עליונה
    refresh_img = load_btn_image("refresh.png")
    tk.Button(root, image=refresh_img, command=refresh_app, bg="white", borderwidth=0).place(
        relx=1.0, rely=0.0, anchor="ne", x=-10, y=10
    )

    root.mainloop()



if __name__ == "__main__":
    main()

